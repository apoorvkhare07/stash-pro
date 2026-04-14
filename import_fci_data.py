"""
Import FCI Common.xlsx data into StashPro for the FCI organization.
Run with: python manage.py shell < import_fci_data.py
"""
import os
import sys
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'stash_pro.settings')
django.setup()

import openpyxl
from datetime import datetime, date
from decimal import Decimal
from django.contrib.auth.models import User
from django.db import transaction
from accounts.models import Organization, UserOrganization
from inventory.models import Lot, Product, Payment
from sales.models import Sale
from expense.models import Expenses

XLSX_PATH = '/Users/apoorvkhare/Downloads/FCI Common.xlsx'

# --- Mapping ---
PRODUCT_TYPE_MAP = {
    'SLR Body': ('Film Camera', 'SLR'),
    'SLR': ('Film Camera', 'SLR'),
    'Lens': ('Accessory', 'Lens'),
    'Point & Shoot': ('Film Camera', 'Point & Shoot'),
    'Digicam': ('Digital Camera', 'Point & Shoot'),
    'Rangefinder': ('Film Camera', 'Rangefinder'),
    'TLR': ('Film Camera', 'TLR'),
    'Mirrorless': ('Digital Camera', 'Mirrorless'),
    'DSLR Body': ('Digital Camera', 'SLR'),
    'Handycam': ('Accessory', 'Handycam'),
    'Flash': ('Accessory', None),
    'Lens Adapter': ('Accessory', None),
    'Tele Converter': ('Accessory', None),
    'Other': ('Accessory', None),
    'Expired film': ('Film', 'Film Roll'),
    'P&S + Film': ('Film Camera', 'Point & Shoot'),
    'Viewfinder': ('Accessory', None),
    'Shutter Release cable': ('Accessory', None),
}

BUYER_MAP = {
    'Jayesh': 'jayesh',
    'Khare': 'admin',  # you
    'FCI': None,  # org-funded
}


def get_or_create_user(username, org):
    user, created = User.objects.get_or_create(
        username=username,
        defaults={'is_active': True}
    )
    if created:
        user.set_password('changeme123')
        user.save()
        print(f'  Created user: {username}')
    UserOrganization.objects.get_or_create(
        user=user, organization=org,
        defaults={'role': 'editor'}
    )
    return user


def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


@transaction.atomic
def run_import():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    org = Organization.objects.get(slug='fci')
    print(f'Importing into org: {org.name} (id={org.id})')

    # Ensure users exist
    user_cache = {}
    for buyer_name, username in BUYER_MAP.items():
        if username:
            user_cache[buyer_name] = get_or_create_user(username, org)

    # ========== INVENTORY ==========
    ws = wb['Inventory']
    lot_cache = {}  # lot_number -> Lot
    product_cache = {}  # product_id -> Product

    # First pass: collect lot info (group by lot number)
    lot_info = {}  # lot_num -> {source, date, total_price, buyer}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        product_id = int(row[0]) if row[0] else None
        lot_num = int(row[1]) if row[1] else None
        buyer = row[3]
        name = row[4]
        purchase_date = parse_date(row[5])
        source = row[6]
        buying_price = Decimal(str(row[7])) if row[7] else Decimal('0')

        if lot_num and lot_num not in lot_info:
            lot_info[lot_num] = {
                'source': source,
                'date': purchase_date,
                'total_price': Decimal('0'),
                'buyer': buyer,
                'products': [],
            }
        if lot_num:
            lot_info[lot_num]['total_price'] += buying_price
            # Use the earliest date as lot date
            if purchase_date and (lot_info[lot_num]['date'] is None or purchase_date < lot_info[lot_num]['date']):
                lot_info[lot_num]['date'] = purchase_date

    # Create lots
    for lot_num, info in sorted(lot_info.items()):
        buyer = info['buyer']
        funded_by = 'org' if buyer == 'FCI' else 'user'
        funded_by_user = user_cache.get(buyer)

        lot = Lot.objects.create(
            organization=org,
            title=f"Lot #{lot_num}",
            total_price=info['total_price'],
            bought_on=info['date'] or date(2026, 3, 1),
            bought_from=info['source'] or '',
            status='paid',
            funded_by=funded_by,
            funded_by_user=funded_by_user if funded_by == 'user' else None,
        )
        lot_cache[lot_num] = lot
        print(f'  Lot #{lot_num}: {info["source"]} - ₹{info["total_price"]} (funded_by={funded_by})')

    # Second pass: create products
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        product_id = int(row[0])
        lot_num = int(row[1]) if row[1] else None
        product_type = row[2]
        name = row[4]
        purchase_date = parse_date(row[5])
        source = row[6]
        buying_price = Decimal(str(row[7])) if row[7] else Decimal('0')
        remarks = row[8]
        selling_price = Decimal(str(row[9])) if row[9] else None
        sold = str(row[12]).strip().lower() == 'yes' if row[12] else False
        status = row[13]  # Listed, Unlisted, Shipped

        category, sub_category = PRODUCT_TYPE_MAP.get(product_type, ('Accessory', None))

        # Determine available quantity
        available = 0 if sold else 1

        # Map status to delivery_status
        delivery_status = 'received'

        product = Product.objects.create(
            organization=org,
            lot=lot_cache.get(lot_num),
            name=name or f'Product #{product_id}',
            price=buying_price,
            stock=1,
            available_quantity=available,
            category=category,
            sub_category=sub_category,
            bought_from=source,
            bought_at=purchase_date,
            delivery_status=delivery_status,
            overall_condition=remarks or '',
        )
        product_cache[product_id] = product

    print(f'  Created {len(product_cache)} products')

    # ========== SALES ==========
    ws = wb['Sales']
    sales_created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        order_id = str(row[0])  # e.g. "#1201"
        sale_date = row[1]
        shopify_name = row[2]
        customer_email = row[3]
        sale_price = Decimal(str(row[4])) if row[4] else Decimal('0')
        product_id_1 = int(row[5]) if row[5] else None
        product_id_2 = int(row[6]) if row[6] else None
        cost_price = Decimal(str(row[12])) if row[12] else None

        # Create sale for product 1
        product = product_cache.get(product_id_1)
        if not product:
            print(f'  WARN: Product {product_id_1} not found for sale {order_id}')
            continue

        # Determine funded_by_user from product's lot
        funded_by_user = None
        if product.lot and product.lot.funded_by == 'user':
            funded_by_user = product.lot.funded_by_user

        # For multi-product orders, split price evenly (or use cost as guide)
        items_in_order = 1 + (1 if product_id_2 else 0)

        sale = Sale.objects.create(
            organization=org,
            product=product,
            quantity_sold=1,
            sale_price=sale_price if items_in_order == 1 else sale_price,  # full price for single, will handle multi below
            customer=customer_email or '',
            sale_date=sale_date,
            shopify_order_id=order_id,
            shopify_order_name=order_id,
            shipping_status='shipped',
            funded_by_user=funded_by_user,
            cost_price=product.price,
        )
        sale.calculate_split()
        sale.save()
        sales_created += 1

        # Product 2 if exists (multi-product order)
        if product_id_2:
            product2 = product_cache.get(product_id_2)
            if product2:
                funded_by_user2 = None
                if product2.lot and product2.lot.funded_by == 'user':
                    funded_by_user2 = product2.lot.funded_by_user

                sale2 = Sale.objects.create(
                    organization=org,
                    product=product2,
                    quantity_sold=1,
                    sale_price=Decimal('0'),  # price is on the main order
                    customer=customer_email or '',
                    sale_date=sale_date,
                    shopify_order_id=f'{order_id}-2',  # suffix for uniqueness
                    shopify_order_name=order_id,
                    shipping_status='shipped',
                    funded_by_user=funded_by_user2,
                    cost_price=product2.price,
                )
                sale2.calculate_split()
                sale2.save()
                sales_created += 1

    print(f'  Created {sales_created} sales')

    # ========== EXPENSES ==========
    ws = wb['Expenses']
    expenses_created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        exp_date = parse_date(row[1])
        exp_type = row[2]
        description = row[3]
        spent_by = row[4]
        amount = Decimal(str(row[5])) if row[5] else Decimal('0')

        # Map expense type
        expense_type = 'misc'
        if description:
            desc_lower = description.lower()
            if 'delhivery' in desc_lower or 'shipping' in desc_lower or 'courier' in desc_lower:
                expense_type = 'shipping'
            elif 'service' in desc_lower or 'repair' in desc_lower or 'battery' in desc_lower:
                expense_type = 'servicing'

        Expenses.objects.create(
            organization=org,
            type=expense_type,
            amount=amount,
            description=description or '',
            vendor=spent_by or '',
            date=exp_date or date(2026, 4, 1),
        )
        expenses_created += 1

    print(f'  Created {expenses_created} expenses')
    print('\nDone!')


if __name__ == '__main__':
    run_import()
else:
    run_import()
